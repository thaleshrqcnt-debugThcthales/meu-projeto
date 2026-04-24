"""
Exporta orçamento para XLSX usando openpyxl.
"""
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
TOTAL_FILL = PatternFill(start_color="e8f4f8", end_color="e8f4f8", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

THIN = Side(border_style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CATEGORIES = ["recursos_humanos", "servicos", "materiais", "comunicacao", "acessibilidade", "registro", "outros"]


def _header_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    return cell


def _data_cell(ws, row, col, value, bold=False, number_format=None, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", size=10, bold=bold)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = BORDER
    if number_format:
        cell.number_format = number_format
    if fill:
        cell.fill = fill
    return cell


def export_budget_xlsx(project, budgets: list, output_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Orçamento"

    # Título
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"ORÇAMENTO — {project.public_title or project.internal_name}"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="1a1a2e")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    sub_cell = ws["A2"]
    sub_cell.value = (
        f"Proponente: {project.proponent_name} | "
        f"Módulo: {project.selected_module or 'não definido'} | "
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y')}"
    )
    sub_cell.font = Font(name="Calibri", size=10, italic=True)
    sub_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Cabeçalho
    headers = ["#", "Item", "Categoria", "Descrição", "Unid.", "Qtd.", "Valor Unit. (R$)", "Valor Total (R$)", "Justificativa / Relação com Ação"]
    for col, h in enumerate(headers, 1):
        _header_cell(ws, 3, col, h)
    ws.row_dimensions[3].height = 35

    # Dados
    total = 0.0
    category_totals: dict[str, float] = {}

    for idx, b in enumerate(budgets, 1):
        row_num = idx + 3
        cat = (b.category or "outros").lower()
        total_val = b.quantity * b.unit_value
        total += total_val
        category_totals[cat] = category_totals.get(cat, 0) + total_val

        _data_cell(ws, row_num, 1, idx)
        _data_cell(ws, row_num, 2, b.item or "")
        _data_cell(ws, row_num, 3, b.category or "")
        _data_cell(ws, row_num, 4, b.description or "")
        _data_cell(ws, row_num, 5, b.unit or "")
        _data_cell(ws, row_num, 6, b.quantity, number_format="#,##0.##")
        _data_cell(ws, row_num, 7, b.unit_value, number_format='R$ #,##0.00')
        _data_cell(ws, row_num, 8, total_val, number_format='R$ #,##0.00')
        _data_cell(ws, row_num, 9, b.justification or "")

        # Alerta se acessibilidade zerada
        if "acessibilidade" in cat and total_val == 0:
            ws.cell(row=row_num, column=8).fill = ALERT_FILL

    # Linha de total
    total_row = len(budgets) + 4
    ws.merge_cells(f"A{total_row}:G{total_row}")
    total_label = ws.cell(row=total_row, column=1, value="TOTAL GERAL")
    total_label.font = Font(name="Calibri", bold=True, size=11)
    total_label.alignment = Alignment(horizontal="right")
    total_label.fill = TOTAL_FILL

    total_cell = ws.cell(row=total_row, column=8, value=total)
    total_cell.font = Font(name="Calibri", bold=True, size=11)
    total_cell.number_format = 'R$ #,##0.00'
    total_cell.fill = TOTAL_FILL
    total_cell.border = BORDER

    # Alerta se excede módulo
    intended = project.intended_budget or 0
    if intended > 0 and total > intended * 1.001:
        ws.cell(row=total_row + 1, column=1).value = f"⚠ ATENÇÃO: Total (R$ {total:,.2f}) excede o valor do módulo (R$ {intended:,.2f})"
        ws.cell(row=total_row + 1, column=1).font = Font(name="Calibri", bold=True, color="CC0000")
        ws.merge_cells(f"A{total_row + 1}:I{total_row + 1}")

    # Aba de resumo por categoria
    ws2 = wb.create_sheet("Resumo por Categoria")
    _header_cell(ws2, 1, 1, "Categoria")
    _header_cell(ws2, 1, 2, "Valor (R$)")
    _header_cell(ws2, 1, 3, "% do Total")

    for i, cat in enumerate(CATEGORIES, 2):
        val = category_totals.get(cat, 0)
        pct = (val / total * 100) if total > 0 else 0
        _data_cell(ws2, i, 1, cat.replace("_", " ").title())
        _data_cell(ws2, i, 2, val, number_format='R$ #,##0.00')
        _data_cell(ws2, i, 3, pct / 100, number_format="0.0%")

        # Alerta acessibilidade = 0
        if cat == "acessibilidade" and val == 0:
            ws2.cell(row=i, column=2).fill = ALERT_FILL
            ws2.cell(row=i, column=3).fill = ALERT_FILL

        # Alerta comunicação = 0
        if cat == "comunicacao" and val == 0:
            ws2.cell(row=i, column=2).fill = WARN_FILL
            ws2.cell(row=i, column=3).fill = WARN_FILL

    # Larguras das colunas (aba principal)
    col_widths = [5, 30, 18, 35, 8, 8, 16, 16, 45]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 12

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return str(path)
